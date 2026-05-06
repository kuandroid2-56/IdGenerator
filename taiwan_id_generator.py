"""
Taiwan-style ID number generator.
Format: [A-Z][1-2][0-9]{7}[check_digit]
Usage: python3 taiwan_id_generator.py <count> [output_filename]
"""

import random
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Region letter to two-digit code mapping (Taiwan standard)
LETTER_CODES = {
    'A': 10, 'B': 11, 'C': 12, 'D': 13, 'E': 14, 'F': 15,
    'G': 16, 'H': 17, 'I': 34, 'J': 18, 'K': 19, 'L': 20,
    'M': 21, 'N': 22, 'O': 35, 'P': 23, 'Q': 24, 'R': 25,
    'S': 26, 'T': 27, 'U': 28, 'V': 29, 'W': 32, 'X': 30,
    'Y': 31, 'Z': 33,
}

# Multipliers for each position (positions 0-9 after expanding the letter)
MULTIPLIERS = [1, 9, 8, 7, 6, 5, 4, 3, 2, 1]


def calc_check_digit(letter: str, gender: int, digits: list[int]) -> int:
    """Calculate the check digit for a Taiwan ID number."""
    code = LETTER_CODES[letter]
    d1 = code // 10  # tens digit of region code
    d2 = code % 10   # units digit of region code
    values = [d1, d2, gender] + digits
    total = sum(v * m for v, m in zip(values, MULTIPLIERS))
    return (10 - (total % 10)) % 10


def generate_id() -> str:
    """Generate a single valid Taiwan-style ID number."""
    letter = random.choice(list(LETTER_CODES.keys()))
    gender = random.choice([1, 2])
    digits = [random.randint(0, 9) for _ in range(7)]
    check = calc_check_digit(letter, gender, digits)
    return f"{letter}{gender}{''.join(map(str, digits))}{check}"


def generate_ids(count: int) -> list[str]:
    """Generate a list of unique Taiwan-style ID numbers."""
    ids: set[str] = set()
    while len(ids) < count:
        ids.add(generate_id())
    return list(ids)


def save_to_excel(ids: list[str], filename: str) -> None:
    """Save the ID list to an Excel file with 序號 and 身分證號碼 columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "身分證號碼"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20

    for col, title in enumerate(["序號", "身分證號碼"], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for i, id_num in enumerate(ids, start=1):
        fill = alt_fill if i % 2 == 0 else None
        for col, val in enumerate([i, id_num], start=1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = center
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    wb.save(filename)


def main():
    if len(sys.argv) < 2:
        print("用法: python3 taiwan_id_generator.py <數量> [輸出檔名]")
        print("例如: python3 taiwan_id_generator.py 100 output.xlsx")
        sys.exit(1)

    try:
        count = int(sys.argv[1])
        if count <= 0:
            raise ValueError
    except ValueError:
        print("錯誤：數量必須為正整數")
        sys.exit(1)

    filename = sys.argv[2] if len(sys.argv) >= 3 else "taiwan_ids.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    print(f"正在產生 {count} 筆台灣身分證號碼格式...")
    ids = generate_ids(count)
    save_to_excel(ids, filename)
    print(f"已儲存至 {filename}")


if __name__ == "__main__":
    main()
