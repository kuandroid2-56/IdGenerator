"""
De-identification tool for Taiwan-style ID numbers.
Reads an Excel file produced by taiwan_id_generator.py, appends a
去識別化身分證號碼 column using HMAC-SHA256, and saves a new file.

Why HMAC-SHA256 instead of simple masking:
  - Masking (e.g. A12****789) causes collisions: up to 10,000 different original
    IDs share the same masked value, breaking 1-to-1 correspondence.
  - Plain SHA-256 is collision-free in practice but vulnerable to brute-force
    because the Taiwan ID keyspace is finite (~520 million valid IDs).
  - HMAC-SHA256 with a secret key makes brute-force computationally infeasible
    without the key, while still guaranteeing a unique token per unique ID.

Secret key handling:
  - Pass via --key <secret> argument, or set env var DEIDENTIFY_KEY.
  - If neither is provided a random key is generated, printed, and must be
    stored to re-produce the same tokens later.

Usage:
  python3 taiwan_id_deidentify.py <input.xlsx> [output.xlsx] [--key <secret>]
"""

import hashlib
import hmac
import os
import secrets
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def hmac_token(id_str: str, key: bytes) -> str:
    """Return a 16-char uppercase hex HMAC-SHA256 token for id_str."""
    digest = hmac.new(key, id_str.encode(), hashlib.sha256).hexdigest()
    return digest[:16].upper()


def process_excel(input_path: str, output_path: str, key: bytes) -> None:
    try:
        src = load_workbook(input_path)
    except FileNotFoundError:
        print(f"錯誤：找不到檔案 '{input_path}'")
        sys.exit(1)
    except Exception as e:
        print(f"錯誤：無法開啟檔案 – {e}")
        sys.exit(1)

    ws_src = src.active
    rows = list(ws_src.iter_rows(values_only=True))

    if not rows:
        print("錯誤：來源檔案沒有資料")
        sys.exit(1)

    header = list(rows[0])
    data = rows[1:]

    id_col_index = None
    for i, col_name in enumerate(header):
        if col_name == "身分證號碼":
            id_col_index = i
            break

    if id_col_index is None:
        print("錯誤：找不到「身分證號碼」欄位，請確認輸入檔案格式正確")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "去識別化身分證號碼"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    new_col_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    alt_new_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    new_header = header + ["去識別化身分證號碼"]
    col_widths = [12] * len(header) + [22]

    for col_idx, (title, width) in enumerate(zip(new_header, col_widths), start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = new_col_fill if col_idx == len(new_header) else header_fill
        cell.alignment = center

    for row_idx, row in enumerate(data, start=2):
        is_alt = row_idx % 2 == 0
        id_value = row[id_col_index] if id_col_index < len(row) else ""
        token = hmac_token(str(id_value), key) if id_value else ""

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center
            if is_alt:
                cell.fill = alt_fill

        new_col_idx = len(row) + 1
        cell = ws.cell(row=row_idx, column=new_col_idx, value=token)
        cell.alignment = center
        if is_alt:
            cell.fill = alt_new_fill

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"完成：共處理 {len(data)} 筆，已儲存至 {output_path}")
    print(f"去識別化方式：HMAC-SHA256（取前 16 碼十六進位）")
    print(f"範例：A123456789  →  3D9F2A... (唯一、不可逆)")


def parse_args() -> tuple[str, str, bytes]:
    args = sys.argv[1:]
    if not args:
        print("用法: python3 taiwan_id_deidentify.py <輸入檔案.xlsx> [輸出檔案.xlsx] [--key <密鑰>]")
        print("例如: python3 taiwan_id_deidentify.py taiwan_ids.xlsx deidentified.xlsx --key mysecret")
        sys.exit(1)

    input_path = args[0]
    output_path = None
    key_str = None

    i = 1
    while i < len(args):
        if args[i] == "--key" and i + 1 < len(args):
            key_str = args[i + 1]
            i += 2
        elif output_path is None and not args[i].startswith("--"):
            output_path = args[i]
            i += 1
        else:
            i += 1

    if output_path is None:
        base = input_path.removesuffix(".xlsx")
        output_path = f"{base}_deidentified.xlsx"
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    # Key priority: --key arg > env var > auto-generate
    if key_str is None:
        key_str = os.environ.get("DEIDENTIFY_KEY")
    if key_str is None:
        key_str = secrets.token_hex(32)
        print(f"[警告] 未提供密鑰，本次自動產生：{key_str}")
        print(f"       請保存此密鑰，日後需產生相同 token 時須使用相同密鑰。")
    else:
        print(f"[資訊] 使用指定密鑰進行 HMAC-SHA256 去識別化。")

    return input_path, output_path, key_str.encode()


def main():
    input_path, output_path, key = parse_args()
    process_excel(input_path, output_path, key)


if __name__ == "__main__":
    main()
